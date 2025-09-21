"""DailyGammon Score Synchronizer
------------------------------

This script synchronizes tournament match results between an Excel results table
and DailyGammon (DG). It automates the process of filling in missing match IDs,
fetching match results, and updating scores into the correct table cells.

This script processes match results for a specific league and writes them to Excel. 
This script can also run across multiple leagues when used together with the wrapper script.


Usage:
    - Manual mode (default):
        Simply run the script without arguments. 
        Example: python dailygammon.py
        -> Uses default league "4d" hardwired in the script
        -> Keeps Excel workbook open for manual review

    - Command line / wrapper mode:
        Provide the league as the first argument and optionally '--auto' as the second.
        Example: python dailygammon.py 2b --auto
        -> Processes league "2b"
        -> Closes Excel workbook automatically (needed when running multiple leagues in sequence)

This makes it possible to run the script across multiple leagues 
without changing the source code manually.


Core Concepts:
--------------

1. Match ID Handling
   - If a match_id cell is empty, the script searches DG for invitations
     initiated by the "player" and inserts the found ID into Excel.
   - If a match_id cell is filled, it may be either:
        a) Automatically inserted earlier (normal case)
        b) Manually entered by a moderator (manual ID)
   - Manual IDs are detected by comparing player/opponent order between Excel
     and DG: if reversed, the entry is considered manual.

2. Manual Match IDs
   - Stored separately in `matches_by_hand`.
   - They carry a `switched=True` flag, meaning that for DG lookups the roles
     of player and opponent must be swapped to retrieve results.
   - When writing scores back to Excel, the swapped results are re-switched
     so the table remains consistent from the perspective of the Excel player.

3. Caching
   - Each match_id is requested from DG at most once.
   - A simple dict (`html_cache`) maps { match_id -> html } to reduce load.

4. Idempotence
   - Running the script multiple times does not duplicate work.
   - IDs are inserted only if cells are empty; scores are written only if
     the cell does not already contain a final result (e.g., "11").

5. Score Writing
   - For each resolved match, the correct Excel row and columns are located
     via player/opponent name mapping.
   - Exact (case-insensitive) name matches are preferred.
   - If no exact match is found, a heuristic rule is applied:
       * Check whether one name appears as a substring of the other.
   - If the heuristic is inconclusive, the match is skipped for safety.

6. Safety Rules
   - The script never overwrites an existing score of 11.
   - If names cannot be reliably mapped, the match is skipped instead of
     risking a wrong write.

"""


# ============================================================
# Script Purpose:
# This script automatically updates match results for a DailyGammon league season.
# It connects to DailyGammon with your login credentials, collects all match IDs,
# downloads intermediate/final scores, and writes them into the Excel results file.
#
# Workflow in summary:
#   1. Login to DailyGammon with your credentials
#   2. Read the player list from the Excel "Players" sheet
#   3. Detect already known matches from the "Links" sheet
#   4. Find and insert missing match IDs automatically
#   5. Update "Matches" sheet with intermediate results
#   6. For finished matches, set the final winner score to 11
#
# Excel file requirement:
# - Requires Excel file "<season>th_Backgammon-championships_<league>.xlsm"
#   The corresponding Excel file (e.g. "34th_Backgammon-championships_4d.xlsm")
#   must be located in the same folder as this script.
#
# - Excel sheets used:
#       * "Players" → base player list
#       * "Links"   → references to match IDs
#       * "Matches" → current scores
# - Important: Scores are only updated if the match is not yet marked as finished (11).
#
# Before running, configure:
#   - Your User ID and Password (variables: payload["login"], payload["password"])
#   - Current Season number (variable: saison_nummer, e.g. "34")
#   - League (variable: liga, e.g. "4d")
#
# Required Python libraries:
#   requests, beautifulsoup4, xlwings
#
# If not installed, run:
#   pip install requests beautifulsoup4 xlwings
#
# ============================================================
# ============================================================
# Streamlit-ready Script 1 (OpenPyXL / Pandas, secure login)
# ============================================================
def main():
    import os
    import sys
    import re
    import requests
    from bs4 import BeautifulSoup
    import openpyxl
    import pandas as pd
    from dotenv import load_dotenv
    from datetime import datetime
    import streamlit as st
    import io


    # -----------------------------
    # Load credentials from .env
    # -----------------------------
    load_dotenv(dotenv_path="a.env")
    DG_LOGIN = os.getenv("DG_LOGIN", "")
    DG_PW = os.getenv("DG_PW", "")

    login_url = "http://dailygammon.com/bg/login"
    payload = {
        "login": DG_LOGIN,
        "password": DG_PW,
        "save": "1"
    }

    BASE_URL = "http://dailygammon.com/bg/game/{}/0/list"

    # -----------------------------
    # Season & League Selection
    # -----------------------------
    saison_nummer = "34"  # default season
    if len(sys.argv) > 1:
        liga = sys.argv[1]
    else:
        liga = "4d"  # default league

    AUTO_MODE = "--auto" in sys.argv

    file = f"{saison_nummer}th_Backgammon-championships_{liga}.xlsx"
    output_file = f"{saison_nummer}th_Backgammon-championships_{liga}_output.xlsx"
    season = f"{saison_nummer}th-season-{liga}"

    print("="*50)
    print(f"▶ Script started – collecting links and data for {season}")
    print(f"📂 Excel file: {file}")
    print("="*50)

    # -----------------------------
    # Read players from Excel
    # -----------------------------
    wb_xw = openpyxl.load_workbook(file, data_only=True)
    ws_players = wb_xw["Players"]

    players = []
    player_ids = {}
    for row in ws_players.iter_rows(min_row=2, max_col=1, values_only=False):
        cell = row[0]
        if cell.value:
            name = str(cell.value).strip()
            players.append(name)
            if cell.hyperlink:
                url = cell.hyperlink.target
                player_id = url.rsplit("/", 1)[-1]
                player_ids[name] = player_id
    wb_xw.close()

    # -----------------------------
    # Secure login session
    # -----------------------------
    def login_session() -> requests.Session:
        s = requests.Session()
        s.headers.update({"User-Agent": "Mozilla/5.0"})
        resp = s.post(login_url, data=payload, timeout=30)
        resp.raise_for_status()
        return s

    session = login_session()


    # -----------------------------------------------------
    # --- Collect matches per player ---
    # -----------------------------------------------------
    # -----------------------------------------------------
    # Function: get_player_matches
    # Purpose:
    #   Collects all matches for a specific player in the
    #   given season. It scrapes the DailyGammon user page
    #   and extracts:
    #     - Opponent name
    #     - Opponent ID
    #     - Match ID
    #
    # - Filters table rows by the 'season' string to avoid pulling old matches.
    # -----------------------------------------------------

    def get_player_matches(session: requests.Session, player_id, season):
        url = f"http://www.dailygammon.com/bg/user/{player_id}"
        r = session.get(url)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        player_matches = []
        for row in soup.find_all("tr"):
            text = row.get_text(" ", strip=True)
            if season not in text:
                continue
            opponent_link = row.find("a", href=re.compile(r"/bg/user/\d+"))
            match_link = row.find("a", href=re.compile(r"/bg/game/\d+/0/"))
            if not opponent_link or not match_link:
                continue
            opponent_name = opponent_link.text.strip()
            opponent_id = re.search(r"/bg/user/(\d+)", opponent_link["href"]).group(1)
            match_id = re.search(r"/bg/game/(\d+)/0/", match_link["href"]).group(1)
            player_matches.append((opponent_name, opponent_id, match_id))
        return player_matches

    # -----------------------------------------------------
    # --- Helper functions: fetch HTML & extract scores ---
    # -----------------------------------------------------
    # -----------------------------------------------------
    # Function: fetch_list_html
    # Purpose:
    #   Downloads the HTML page for a specific match ID.
    #   Returns the HTML content or None if the request failed.
    # -----------------------------------------------------

    def fetch_list_html(session: requests.Session, match_id: int) -> str | None:
        url = BASE_URL.format(match_id)
        try:
            resp = session.get(url, timeout=30)
            if not resp.ok or "Please Login" in resp.text:
                return None
            return resp.text
        except requests.RequestException:
            return None

    # -----------------------------------------------------
    # Function: extract_latest_score
    # Purpose:
    #   Parses the match HTML page and extracts the latest
    #   visible score row for the two players.
    #   Returns player names + current scores.
    #
    # PARSE LATEST SCORE FROM MATCH PAGE:
    # - Scans table rows from bottom to top (reversed) to find the most recent score line.
    # - Assumes the pattern "<Name> : <Score>" is present on both left and right columns.
    # -----------------------------------------------------

    def extract_latest_score(html: str, players_list: list[str]):
        soup = BeautifulSoup(html, "html.parser")
        for row in reversed(soup.find_all("tr")):
            text = row.get_text(" ", strip=True)
            if not any(p in text for p in players_list):
                continue
            cells = row.find_all("td")
            if len(cells) >= 3:
                left_text = cells[1].get_text(" ", strip=True)
                right_text = cells[2].get_text(" ", strip=True)
                left_match = re.match(r"(.+?)\s*:\s*(\d+)", left_text)
                right_match = re.match(r"(.+?)\s*:\s*(\d+)", right_text)
                if left_match and right_match:
                    left_name, left_score = left_match.groups()
                    right_name, right_score = right_match.groups()
                    return left_name.strip(), right_name.strip(), int(left_score), int(right_score)
        return None

    # -----------------------------------------------------
    # Function: map_scores_for_excel
    # Purpose:
    #   Aligns scores from DailyGammon with the correct order
    #   in the Excel sheet.
    #   Handles switched cases (player order reversed for manual added matches).
    #
    # NAME/SCORE ALIGNMENT TO EXCEL:
    # - The Excel grid expects "excel_player" vs "excel_opponent" in a fixed orientation.
    # - 'switched_flag=True' means the match was manually entered with reversed order
    #   (excel_player appears on the right on DailyGammon), so we swap scores here.
    # - If names match exactly (case-insensitive), we map directly; otherwise we use a
    #   small heuristic (substring check) as a fallback. If unsure, return None (skip write).
    # -----------------------------------------------------


    def map_scores_for_excel(player, opponent, left_name, right_name, left_score, right_score, switched_flag):
        ln = left_name.strip().lower()
        rn = right_name.strip().lower()
        pn = player.strip().lower()
        on = opponent.strip().lower()

        if switched_flag:
            return right_score, left_score

        if ln == pn and rn == on:
            return left_score, right_score
        if ln == on and rn == pn:
            return right_score, left_score

        # Fallback heuristic if names differ slightly
        if pn in ln or pn in rn or on in ln or on in rn:
            if pn in ln:
                return left_score, right_score
            if pn in rn:
                return right_score, left_score
        return None

    # -----------------------------------------------------
    # --- Open Excel workbook via xlwings ---
    # -----------------------------------------------------
    # EXCEL WRITING PHASE (XLWINGS):
    # - From this point on, we interact with Excel via xlwings (live Excel instance).
    # -----------------------------------------------------

    ws_links = wb_xw["Links"]
    ws_matches = wb_xw["Matches"]

    # Extract players/columns from "Links"
    # "LINKS" SHEET LAYOUT ASSUMPTION:
    # - Column A (from row 2 down) lists row player names.
    # - Row 1 (from column B rightwards) lists opponent names (as columns).
    # - Cells at (row_player, col_opponent) hold the match ID (and hyperlink).

    row_players_links = []
    r = 2
    while (v := ws_links[f"A{r}"].value):
        row_players_links.append(str(v).strip())
        r += 1

    col_opponents_links = []
    c = 2
    while (v := ws_links.cell(row=1, column=c).value):
        col_opponents_links.append(str(v).strip())
        c += 1

    col_index_links = {name: 2 + i for i, name in enumerate(col_opponents_links)}

    print("DEBUG: row_players_links =", row_players_links)
    print("DEBUG: col_opponents_links =", col_opponents_links)
    print("DEBUG: col_index_links =", col_index_links)


    # create WS Match_flag&control
    if "match_flag" in wb_xw.sheetnames:
        ws_flag = wb_xw["match_flag"]
    else:
        ws_flag = wb_xw.create_sheet("match_flag")

    # Row 1: Opponent names
    for col_idx, opp_name in enumerate(col_opponents_links, start=2):
        ws_flag.cell(row=1, column=col_idx).value = opp_name

    # Column A: Player names
    for row_idx, player_name in enumerate(row_players_links, start=2):
        ws_flag.cell(row=row_idx, column=1).value = player_name

    if "control" in wb_xw.sheetnames:
        ws_control = wb_xw["control"]
    else:
        ws_control = wb_xw.create_sheet("control")

    #skip fetching and match_id searching

    skip_fetching = False
    if ws_control["A1"].value == "All match IDs filled":
        print("✅ All match IDs already filled. Skipping fetching and Step 1/Step 2.")
        skip_fetching = True



    # -----------------------------------------------------
    # --- Data structures ---
    # -----------------------------------------------------
    matches = {}
    matches_by_hand = {}
    match_id_to_excel = {}
    html_cache = {}
    finished_by_id = {}


    #Skippimg Step 1 & 2 because matches are filled
    print(f"DEBUG: ws_control['A1'] = {ws_control['A1'].value!r}")
    print(f"DEBUG: skip_fetching = {skip_fetching}")
    print(f"DEBUG: entering Step 1 & 2? {'Yes' if not skip_fetching else 'No'}")

    if not skip_fetching:


        # -----------------------------------------------------
        # Step 1: Check existing links
        # Purpose:
        #   Go through the "Links" sheet and verify which matches
        #   already have a match ID entered. If the IDs are present,
        #   confirm whether the player/opponent order is correct.
        #   Marks switched matches if detected.
        # -----------------------------------------------------

        for i, player_name in enumerate(row_players_links, start=2):
            for opp in col_opponents_links:
                if player_name == opp:
                    continue
                c = col_index_links.get(opp)
            
                val = ws_links.cell(row=i, column=c).value
                

                if not val:
                    continue
                try:
                    match_id = int(val)
                except Exception:
                    match_id = int(str(val).strip())
                
                # --- Check match_flag sheet to avoid refetching ---
                row_idx_flag = row_players_links.index(player_name) + 2
                col_idx_flag = col_index_links.get(opp)
                flag_val = ws_flag.cell(row=row_idx_flag, column=col_idx_flag).value

                if flag_val is None:
                    # Match not yet processed: fetch from DG
                    html_cache[match_id] = fetch_list_html(session, match_id)
                else:
                    # Already processed (0 or 1), skip fetching
                    html_cache[match_id] = None

                html = html_cache[match_id]
                if not html:
                    matches[(player_name, opp)] = match_id
                    match_id_to_excel[match_id] = (player_name, opp, False)
                    continue        
                score_info = extract_latest_score(html, [player_name, opp])
                if not score_info:
                    matches[(player_name, opp)] = match_id
                    match_id_to_excel[match_id] = (player_name, opp, False)
                    continue
                left_name, right_name, _, _ = score_info
                ln = left_name.lower(); rn = right_name.lower()
                pn = player_name.lower(); on = opp.lower()
                if ln == pn and rn == on:
                    matches[(player_name, opp)] = match_id
                    match_id_to_excel[match_id] = (player_name, opp, False)

        # MANUAL/SWITCHED CASE:
        # - DailyGammon lists "opponent vs player", but Excel expects "player vs opponent".
        # - We record 'switched=True' for this match_id so all later writes swap correctly.

                elif ln == on and rn == pn:
                    matches_by_hand[(player_name, opp)] = (match_id, True)
                    match_id_to_excel[match_id] = (player_name, opp, True)
                    print(f"Found manual inserted match detected: {player_name} vs {opp} with match ID {match_id}.")
                else:
                    matches[(player_name, opp)] = match_id
                    match_id_to_excel[match_id] = (player_name, opp, False)
                    print(f"⚠️ Unclear order for match ID {match_id}: DG shows '{left_name}' vs '{right_name}'")

        # -----------------------------------------------------
        # Step 2: Fill missing match IDs
        # Purpose:
        #   For each player, check which opponents still have no
        #   match ID in the "Links" sheet. Search for the match on
        #   DailyGammon and insert it automatically into the table.
        #   Also detects "switched" matches (player/opponent reversed).
        #
        # STEP 2 RATIONALE:
        # - For any missing (player, opponent) cell, we look up the player's page to find
        #   their active matches for this season and backfill the match ID into "Links".
        # - We also attach a hyperlink to the specific match list page for quick access.
        # - Existing cells are left untouched; only empty cells get filled.
        # -----------------------------------------------------


        for player in players:
            pid = player_ids.get(player)
            if not pid:
                continue
            missing = [opp for opp in players if opp != player and (player, opp) not in matches and (player, opp) not in matches_by_hand]
            if not missing:
                continue
            player_matches = get_player_matches(session, pid, season=season)
            for opponent_name, opponent_id, match_id in player_matches:
                key = (player, opponent_name)
                if key in matches or key in matches_by_hand:
                    continue
                mid_int = int(match_id)
                switched_flag = False
                if mid_int in match_id_to_excel:
                    _, _, switched_flag = match_id_to_excel[mid_int]
                matches[key] = mid_int
                match_id_to_excel[mid_int] = (player, opponent_name, switched_flag)
                try:
                    row_idx = row_players_links.index(player) + 2
                except ValueError:
                    continue
                c = col_index_links.get(opponent_name)
                if not c or opponent_name == player:
                    continue

                cell = ws_links.cell(row=row_idx, column=c)

                # --- match_flag sheet setzen ---
                row_idx_flag = row_players_links.index(player) + 2
                col_idx_flag = col_index_links.get(opponent_name)
                ws_flag.cell(row=row_idx_flag, column=col_idx_flag).value = 1 if switched_flag else 0


        # - We only write if the cell is empty to avoid overwriting manual adjustments.
        # - If you ever need to refresh a wrong ID, clear the cell first, then rerun.

                if not cell.value:
                    cell.value = str(match_id)
                    cell.hyperlink = f"http://www.dailygammon.com/bg/game/{match_id}/0/list#end"
                    print(f"Detected missing match between {player} and {opponent_name} — match ID={match_id} has been auto-added to the table")

        print("✅ Match IDs updated (auto + manual detection)")

    else:
        print("Skipping Step 1 & Step 2: using existing match IDs only.")

    # -----------------------------------------------------
    # Step 3: Collect finished matches
    # Purpose:
    #   For every player, fetch their export page.
    #   If a match is marked as finished, extract the winner.
    #   Results are stored in a dictionary for later processing.
    #
    # FINISHED MATCH DETECTION:
    # - We open each player's page and follow "export" links for matches of this season.
    # - The winner is inferred from a simple textual rule (position of "Wins" on the line).
    # - 'finished_by_id' maps match_id -> winner_name for later use in Phase 2.
    # -----------------------------------------------------

    for player in players:
        pid = player_ids.get(player)
        if not pid:
            continue
        url = f"http://www.dailygammon.com/bg/user/{pid}"
        try:
            r = session.get(url, timeout=30)
            r.raise_for_status()
        except requests.RequestException:
            continue
        soup = BeautifulSoup(r.text, "html.parser")
        for row in soup.find_all("tr"):
            text = row.get_text(" ", strip=True)
            if season not in text:
                continue
            export_link = row.find("a", href=re.compile(r"/bg/export/\d+"))
            match_link = row.find("a", href=re.compile(r"/bg/game/\d+/0/"))
            opponent_link = row.find("a", href=re.compile(r"/bg/user/\d+"))
            if not export_link or not match_link or not opponent_link:
                continue
            try:
                match_id = int(re.search(r"/bg/game/(\d+)/0/", match_link["href"]).group(1))
            except Exception:
                continue
            opponent_name = opponent_link.text.strip()
            export_url = f"http://www.dailygammon.com/bg/export/{match_id}"
            try:
                resp_export = session.get(export_url, timeout=30)
                text_lines = resp_export.text.splitlines()
            except requests.RequestException:
                continue
            winner = None

    # - 'mid_threshold 24' is a rough character-position cutoff to decide whether the "Wins"
    #   belongs to the left or right player on the export line.

            mid_threshold = 24
            for line in text_lines:
                if "and the match" in line and "Wins" in line:
                    pos = line.find("Wins")
                    winner = player if pos < mid_threshold else opponent_name
                    break
            if winner:
                finished_by_id[match_id] = winner

    # -----------------------------------------------------
    # Phase 1: Write intermediate scores
    # Purpose:
    #   For each match, download the latest score and update
    #   the "Matches" sheet in Excel.
    #   IMPORTANT: If a score of 11 is already present,
    #   the match is considered finished and will not be overwritten.
    # -----------------------------------------------------
    print(f"Total matches to process: {len(match_id_to_excel)}")
    for mid, info in match_id_to_excel.items():
        print(f"Match ID {mid}: Excel mapping: {info}, HTML fetched? {'Yes' if html_cache.get(mid) else 'No'}")

    print("🔎 Phase 1: Writing intermediate scores for matches...")
    players_in_matches = []
    row_counter = 4
    while True:
        nm = ws_matches.cell(row=row_counter, column=1).value
        if not nm:
            break
        players_in_matches.append(str(nm).strip())
        row_counter += 1
    col_start = 2

    # Deterministische Zeilen-/Spaltenzuordnung für Matches
    row_index_matches = {name: i+4 for i, name in enumerate(players_in_matches)}
    col_index_matches = {name: col_start + i*2 for i, name in enumerate(players_in_matches)}


    # EXCEL WRITE HELPER (INTERMEDIATE SCORES):
    # - Translates (excel_player, excel_opponent) to row/column indices in "Matches".
    # - For each opponent, we reserve two columns: left=excel_player's score, right=excel_opponent's score.
    # - Safety: if either cell already equals 11, we skip to preserve final results.
    # - Scores are already correctly oriented by 'map_scores_for_excel'; no swapping here.

    def write_score_to_excel(excel_player, excel_opponent, player_score, opponent_score, switched_flag):
        try:
            r_idx = players_in_matches.index(excel_player) + 4
            c_base = players_in_matches.index(excel_opponent)
        except ValueError:
            print(f"⚠️ Player not found in Excel sheet: {excel_player} vs {excel_opponent}")
            return False
        c_left = col_start + c_base * 2
        c_right = c_left + 1

        # Do not overwrite already finished (11) scores!
        # - Once a match is finished (11), intermediate updates must never overwrite that cell.

        left_cell_val = ws_matches.cell(row=r_idx, column=c_left).value
        right_cell_val = ws_matches.cell(row=r_idx, column=c_right).value
        if left_cell_val == 11 or right_cell_val == 11:
            return False

        ws_matches.cell(row=r_idx, column=c_left).value = player_score
        ws_matches.cell(row=r_idx, column=c_right).value = opponent_score

        return True

    # - Pull HTML from cache if available; otherwise fetch fresh.

    for match_id, (excel_player, excel_opponent, switched_flag) in list(match_id_to_excel.items()):
        html = html_cache.get(match_id)
        if not html:
            html = fetch_list_html(session, match_id)
            html_cache[match_id] = html
        if not html:
            continue
        result = extract_latest_score(html, players_in_matches)
        if not result:
            continue
        left_name, right_name, left_score, right_score = result

        # Map scores based on player names

        mapped = map_scores_for_excel(excel_player, excel_opponent, left_name, right_name, left_score, right_score, switched_flag)
        if mapped is None:
            continue
        excel_player_score, excel_opponent_score = mapped
        write_score_to_excel(excel_player, excel_opponent, excel_player_score, excel_opponent_score, switched_flag)

    print("✅ Phase 1: completed")
    # -----------------------------------------------------
    # Phase 2: Final results - Set winners to 11 points
    # Purpose:
    #   For matches identified as finished, write the final
    #   winner score (11 points) into the correct player cell
    #   in the "Matches" sheet.
    # -----------------------------------------------------

    print("🔎 Phase 2: Final results (set winner = 11) ...")
    for match_id, winner_name in finished_by_id.items():
        info = match_id_to_excel.get(match_id)
        if not info:
            continue
        excel_player, excel_opponent, switched_flag = info
        try:
            r_idx = players_in_matches.index(excel_player) + 4
            c_base = players_in_matches.index(excel_opponent)
        except ValueError:
            continue
        c_left = col_start + c_base * 2
        c_right = c_left + 1

        # Write 11 to the correct winner cell

        winner_lower = winner_name.strip().lower()
        if switched_flag:
            if winner_lower == excel_player.lower():
                ws_matches.cell(row=r_idx, column=c_right).value = 11

            elif winner_lower == excel_opponent.lower():
                ws_matches.cell(row=r_idx, column=c_left).value = 11
        else:
            if winner_lower == excel_player.lower():
                ws_matches.cell(row=r_idx, column=c_left).value = 11

            elif winner_lower == excel_opponent.lower():
                ws_matches.cell(row=r_idx, column=c_right).value = 11

    all_filled = True

    # Check Links sheet
    for i, player_name in enumerate(row_players_links, start=2):
        for opp in col_opponents_links:
            if player_name == opp:
                continue
            c = col_index_links.get(opp)
            if not ws_links.cell(row=i, column=c).value:
                all_filled = False
                break
        if not all_filled:
            break

    # Or, equivalently, check match_flag 0/1 values
    # for i in range(2, len(row_players_links)+2):
    #     for j in range(2, len(col_opponents_links)+2):
    #         val = ws_flag.cell(row=i, column=j).value
    #         if val is None:
    #             all_filled = False
    #             break
    #     if not all_filled:
    #         break
    if all_filled:
        ws_control["A1"].value = "All match IDs filled"

    #wb_xw.save(file)
    # Close workbook automatically only if called from wrapper
    #if AUTO_MODE:
    #    wb_xw.close()

    # ============================================================
    # Optional: Persistenz im Streamlit-Cloud-Umfeld
    # ============================================================

    try:

        # Excel-Workbook in Memory speichern
        excel_bytes = io.BytesIO()
        wb_xw.save(excel_bytes)
        excel_bytes.seek(0)

        # Download-Button in Streamlit anzeigen

        st.download_button(
            label=f"📥 Geänderte Datei {file} herunterladen",
            data=excel_bytes,
            file_name=file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


    except ImportError:
        # Wenn Streamlit nicht installiert ist (lokaler Run), einfach normal speichern
        wb_xw.save(file)
        print(f"💾 Excel-Datei lokal gespeichert: {file}")

    print("🏁 Script finished successfully")
    print("="*50)

if __name__ == "__main__":
    main()